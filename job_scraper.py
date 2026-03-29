"""
Scraper automatique d'offres d'alternance Data Engineer
Sources : France Travail API, Welcome to the Jungle, JobTeaser, Indeed, HelloWork
V2 : Mots-clés étendus + tri par date
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import time
import os
import re

# ───────────────────────────────────────────────
# CONFIG — TOUS LES MOTS-CLÉS DU MÉTIER
# ───────────────────────────────────────────────
KEYWORDS = [
    # Français
    "ingénieur data alternance",
    "ingenieur data alternance",
    "ingénieur big data alternance",
    "ingénieur données alternance",
    "architecte data alternance",
    "ingénieur dataops alternance",
    "ingénieur mlops alternance",
    "développeur data alternance",
    "ingénieur plateforme data alternance",
    # Anglais
    "data engineer alternance",
    "big data engineer alternance",
    "cloud data engineer alternance",
    "data architect alternance",
    "dataops engineer alternance",
    "mlops engineer alternance",
    "analytics engineer alternance",
    "data platform engineer alternance",
    "etl engineer alternance",
    "data pipeline engineer alternance",
    # Mixtes
    "data engineer azure alternance",
    "data engineer aws alternance",
    "data engineer gcp alternance",
    "data engineer spark alternance",
    "consultant data engineer alternance",
]

OUTPUT_FILE = "offres_alternance_data_engineer.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xhtml;q=0.9,*/*;q=0.8",
}

# ───────────────────────────────────────────────
# FRANCE TRAVAIL API
# ───────────────────────────────────────────────
def get_france_travail_token(client_id, client_secret):
    url = "https://entreprise.francetravail.fr/connexion/oauth2/access_token"
    params = {"realm": "/partenaire"}
    data = {
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret,
        "scope": "api_offresdemploiv2 o2dsoffre",
    }
    r = requests.post(url, params=params, data=data, timeout=10)
    r.raise_for_status()
    return r.json()["access_token"]


def scrape_france_travail(client_id=None, client_secret=None):
    if not client_id or not client_secret:
        print("⚠️  France Travail : clés API manquantes")
        return []
    try:
        token = get_france_travail_token(client_id, client_secret)
        url = "https://api.francetravail.io/partenaire/offresdemploi/v2/offres/search"
        headers = {"Authorization": f"Bearer {token}", **HEADERS}
        all_offres = []
        for kw in ["data engineer", "ingénieur data", "big data", "architecte data", "dataops", "mlops", "analytics engineer"]:
            params = {"motsCles": kw, "typeContrat": "CA", "range": "0-49"}
            r = requests.get(url, headers=headers, params=params, timeout=15)
            if r.status_code != 200:
                continue
            for o in r.json().get("resultats", []):
                all_offres.append({
                    "titre": o.get("intitule", ""),
                    "entreprise": o.get("entreprise", {}).get("nom", "N/A"),
                    "lieu": o.get("lieuTravail", {}).get("libelle", ""),
                    "date_publication": o.get("dateCreation", "")[:10] if o.get("dateCreation") else "",
                    "contrat": "Alternance",
                    "source": "France Travail",
                    "lien": f"https://candidat.francetravail.fr/offres/recherche/detail/{o.get('id', '')}",
                    "description": o.get("description", "")[:300],
                })
            time.sleep(1)
        print(f"✅ France Travail : {len(all_offres)} offres")
        return all_offres
    except Exception as e:
        print(f"❌ France Travail : {e}")
        return []


# ───────────────────────────────────────────────
# WELCOME TO THE JUNGLE
# ───────────────────────────────────────────────
def scrape_wttj():
    offres = []
    queries = ["data engineer", "ingénieur data", "big data engineer", "dataops", "mlops", "analytics engineer", "data architect"]
    try:
        for query in queries:
            url = "https://www.welcometothejungle.com/fr/jobs"
            params = {"query": query, "contract_type[]": "apprenticeship", "page": 1, "per_page": 20}
            r = requests.get(url, headers=HEADERS, params=params, timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            cards = soup.find_all("li", {"data-testid": re.compile("job-card")}) or soup.find_all("article")
            for card in cards:
                titre_el = card.find(["h3", "h2", "a"])
                entreprise_el = card.find(attrs={"data-testid": re.compile("company")})
                lieu_el = card.find(attrs={"data-testid": re.compile("location")})
                lien_el = card.find("a", href=True)
                offres.append({
                    "titre": titre_el.get_text(strip=True) if titre_el else "N/A",
                    "entreprise": entreprise_el.get_text(strip=True) if entreprise_el else "N/A",
                    "lieu": lieu_el.get_text(strip=True) if lieu_el else "N/A",
                    "date_publication": datetime.now().strftime("%Y-%m-%d"),
                    "contrat": "Alternance",
                    "source": "Welcome to the Jungle",
                    "lien": "https://www.welcometothejungle.com" + lien_el["href"] if lien_el else "N/A",
                    "description": "",
                })
            time.sleep(1)
        print(f"✅ Welcome to the Jungle : {len(offres)} offres")
        return offres
    except Exception as e:
        print(f"❌ WTTJ : {e}")
        return []


# ───────────────────────────────────────────────
# JOBTEASER
# ───────────────────────────────────────────────
def scrape_jobteaser():
    offres = []
    queries = ["data engineer", "ingénieur data", "big data", "dataops", "mlops", "analytics engineer", "architecte data"]
    try:
        for query in queries:
            url = "https://www.jobteaser.com/fr/job-offers"
            params = {"q": query, "contract_types[]": "apprenticeship", "locale": "fr"}
            r = requests.get(url, headers=HEADERS, params=params, timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            cards = soup.find_all("article") or soup.find_all("div", class_=re.compile("job|offer|card", re.I))
            for card in cards[:20]:
                titre_el = card.find(["h2", "h3", "h4"]) or card.find("a")
                entreprise_el = card.find(class_=re.compile("company|employer", re.I))
                lieu_el = card.find(class_=re.compile("location|city|lieu", re.I))
                date_el = card.find(["time"])
                lien_el = card.find("a", href=True)
                titre = titre_el.get_text(strip=True) if titre_el else ""
                if not titre:
                    continue
                offres.append({
                    "titre": titre,
                    "entreprise": entreprise_el.get_text(strip=True) if entreprise_el else "N/A",
                    "lieu": lieu_el.get_text(strip=True) if lieu_el else "France",
                    "date_publication": date_el.get("datetime", datetime.now().strftime("%Y-%m-%d"))[:10] if date_el else datetime.now().strftime("%Y-%m-%d"),
                    "contrat": "Alternance",
                    "source": "JobTeaser",
                    "lien": "https://www.jobteaser.com" + lien_el["href"] if lien_el and lien_el["href"].startswith("/") else (lien_el["href"] if lien_el else "N/A"),
                    "description": "",
                })
            time.sleep(1)
        print(f"✅ JobTeaser : {len(offres)} offres")
        return offres
    except Exception as e:
        print(f"❌ JobTeaser : {e}")
        return []


# ───────────────────────────────────────────────
# INDEED
# ───────────────────────────────────────────────
def scrape_indeed():
    offres = []
    queries = [
        "data engineer alternance", "ingénieur data alternance",
        "big data engineer alternance", "dataops alternance",
        "mlops alternance", "analytics engineer alternance",
        "architecte data alternance", "etl engineer alternance",
    ]
    try:
        for query in queries:
            url = "https://fr.indeed.com/jobs"
            params = {"q": query, "l": "France", "sort": "date"}
            r = requests.get(url, headers=HEADERS, params=params, timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            cards = soup.find_all("div", class_=re.compile("job_seen_beacon|jobCard|result", re.I))
            for card in cards[:15]:
                titre_el = card.find("h2", class_=re.compile("jobTitle|title", re.I))
                entreprise_el = card.find("span", class_=re.compile("company|employer", re.I))
                lieu_el = card.find("div", class_=re.compile("location|companyLocation", re.I))
                date_el = card.find("span", class_=re.compile("date|posted", re.I))
                lien_el = card.find("a", href=True)
                titre = titre_el.get_text(strip=True) if titre_el else ""
                if not titre:
                    continue
                lien = ""
                if lien_el:
                    href = lien_el.get("href", "")
                    lien = "https://fr.indeed.com" + href if href.startswith("/") else href
                offres.append({
                    "titre": titre,
                    "entreprise": entreprise_el.get_text(strip=True) if entreprise_el else "N/A",
                    "lieu": lieu_el.get_text(strip=True) if lieu_el else "France",
                    "date_publication": date_el.get_text(strip=True) if date_el else datetime.now().strftime("%Y-%m-%d"),
                    "contrat": "Alternance",
                    "source": "Indeed",
                    "lien": lien or "N/A",
                    "description": "",
                })
            time.sleep(1)
        print(f"✅ Indeed : {len(offres)} offres")
        return offres
    except Exception as e:
        print(f"❌ Indeed : {e}")
        return []


# ───────────────────────────────────────────────
# HELLOWORK
# ───────────────────────────────────────────────
def scrape_hellowork():
    offres = []
    queries = [
        "data engineer", "ingénieur data", "big data engineer",
        "dataops", "mlops", "analytics engineer", "architecte data", "etl engineer",
    ]
    try:
        for query in queries:
            url = "https://www.hellowork.com/fr-fr/emploi/recherche.html"
            params = {"k": f"{query} alternance", "c": "alternance", "s": "date"}
            r = requests.get(url, headers=HEADERS, params=params, timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            cards = soup.find_all("article") or soup.find_all("li", class_=re.compile("job|offer", re.I))
            for card in cards[:15]:
                titre_el = card.find(["h2", "h3"])
                entreprise_el = card.find(class_=re.compile("company|employer|entreprise", re.I))
                lieu_el = card.find(class_=re.compile("location|localisation|city", re.I))
                date_el = card.find(["time"])
                lien_el = card.find("a", href=True)
                titre = titre_el.get_text(strip=True) if titre_el else ""
                if not titre:
                    continue
                offres.append({
                    "titre": titre,
                    "entreprise": entreprise_el.get_text(strip=True) if entreprise_el else "N/A",
                    "lieu": lieu_el.get_text(strip=True) if lieu_el else "France",
                    "date_publication": date_el.get("datetime", datetime.now().strftime("%Y-%m-%d"))[:10] if date_el else datetime.now().strftime("%Y-%m-%d"),
                    "contrat": "Alternance",
                    "source": "HelloWork",
                    "lien": lien_el["href"] if lien_el else "N/A",
                    "description": "",
                })
            time.sleep(1)
        print(f"✅ HelloWork : {len(offres)} offres")
        return offres
    except Exception as e:
        print(f"❌ HelloWork : {e}")
        return []


# ───────────────────────────────────────────────
# DÉDUPLICATION
# ───────────────────────────────────────────────
def deduplicate(offres):
    seen = set()
    unique = []
    for o in offres:
        key = (o["titre"].lower().strip(), o["entreprise"].lower().strip())
        if key not in seen:
            seen.add(key)
            unique.append(o)
    return unique


# ───────────────────────────────────────────────
# TRI PAR DATE (du plus récent au plus ancien)
# ───────────────────────────────────────────────
def parse_date(date_str):
    """Convertit une date string en objet datetime pour le tri."""
    if not date_str:
        return datetime.min
    # Formats possibles
    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S", "%d-%m-%Y"]:
        try:
            return datetime.strptime(date_str[:10], fmt)
        except:
            continue
    return datetime.min


def sort_by_date(offres):
    """Trie les offres du plus récent au plus ancien."""
    return sorted(offres, key=lambda o: parse_date(o.get("date_publication", "")), reverse=True)


# ───────────────────────────────────────────────
# EXCEL
# ───────────────────────────────────────────────
def save_to_excel(offres, filepath):
    existing_liens = set()

    if os.path.exists(filepath):
        wb_exist = openpyxl.load_workbook(filepath)
        ws_exist = wb_exist.active
        for row in ws_exist.iter_rows(min_row=2, values_only=True):
            if row and row[6]:
                existing_liens.add(row[6])
        wb_exist.close()

    nouvelles = [o for o in offres if o["lien"] not in existing_liens]
    print(f"\n📊 {len(nouvelles)} nouvelles offres à ajouter")

    if not nouvelles and os.path.exists(filepath):
        print("ℹ️  Aucune nouvelle offre.")
        return

    # Charge ou crée le workbook
    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        # Récupère les offres existantes
        existing_offres = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and any(row):
                existing_offres.append({
                    "titre": row[1] or "",
                    "entreprise": row[2] or "",
                    "lieu": row[3] or "",
                    "date_publication": str(row[4]) if row[4] else "",
                    "contrat": row[5] or "",
                    "lien": row[6] or "",
                    "source": row[7] or "",
                    "date_collecte": str(row[8]) if row[8] else "",
                    "description": row[9] or "",
                })
        # Fusionne et retrie tout
        all_offres_combined = nouvelles + existing_offres
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Offres Alternance"
        all_offres_combined = nouvelles

    # Retrie tout par date
    all_offres_combined = sort_by_date(all_offres_combined)

    # Recrée la feuille proprement
    ws.delete_rows(1, ws.max_row)
    _create_header(ws)

    source_colors = {
        "France Travail": "E8F5E9",
        "Welcome to the Jungle": "E3F2FD",
        "JobTeaser": "FFF3E0",
        "Indeed": "F3E5F5",
        "HelloWork": "FCE4EC",
    }
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, o in enumerate(all_offres_combined, 1):
        row_idx = i + 1
        color = source_colors.get(o.get("source", ""), "FFFFFF")
        fill = PatternFill("solid", start_color=color, end_color=color)

        date_collecte = o.get("date_collecte", datetime.now().strftime("%Y-%m-%d %H:%M"))

        values = [
            i,
            o.get("titre", ""),
            o.get("entreprise", ""),
            o.get("lieu", ""),
            o.get("date_publication", ""),
            o.get("contrat", "Alternance"),
            o.get("lien", ""),
            o.get("source", ""),
            date_collecte,
            o.get("description", ""),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 10))
            cell.font = Font(name="Arial", size=10)

        if o.get("lien") and o.get("lien") != "N/A":
            ws.cell(row=row_idx, column=7).hyperlink = o["lien"]
            ws.cell(row=row_idx, column=7).font = Font(name="Arial", size=10, color="0563C1", underline="single")

    _update_stats_sheet(wb, all_offres_combined)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
    print(f"✅ Fichier sauvegardé : {filepath}")


def _create_header(ws):
    headers = ["#", "Titre du poste", "Entreprise", "Lieu", "Date publication",
               "Type contrat", "Lien", "Source", "Récupéré le", "Description"]
    header_fill = PatternFill("solid", start_color="1565C0", end_color="1565C0")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    col_widths = [5, 40, 25, 20, 18, 15, 50, 20, 18, 60]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 30


def _update_stats_sheet(wb, offres):
    if "Statistiques" in wb.sheetnames:
        ws_stats = wb["Statistiques"]
        ws_stats.delete_rows(1, ws_stats.max_row)
    else:
        ws_stats = wb.create_sheet("Statistiques")

    header_fill = PatternFill("solid", start_color="1565C0", end_color="1565C0")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws_stats["A1"] = "Tableau de bord — Alternances Data Engineer"
    ws_stats["A1"].font = Font(name="Arial", bold=True, size=14, color="1565C0")
    ws_stats.merge_cells("A1:C1")
    ws_stats["A3"] = "Dernière mise à jour :"
    ws_stats["B3"] = datetime.now().strftime("%d/%m/%Y à %H:%M")
    ws_stats["A3"].font = Font(bold=True, name="Arial")
    ws_stats["B3"].font = Font(name="Arial")

    ws_stats["A5"] = "Source"
    ws_stats["B5"] = "Nb offres"
    ws_stats["C5"] = "% du total"
    for cell_ref in ["A5", "B5", "C5"]:
        c = ws_stats[cell_ref]
        c.fill = header_fill
        c.font = Font(bold=True, color="FFFFFF", name="Arial")
        c.alignment = Alignment(horizontal="center")
        c.border = border

    from collections import Counter
    counts = Counter(o.get("source", "Autre") for o in offres)
    row = 6
    for source, count in sorted(counts.items()):
        ws_stats.cell(row=row, column=1, value=source).font = Font(name="Arial")
        ws_stats.cell(row=row, column=1).border = border
        ws_stats.cell(row=row, column=2, value=count).font = Font(name="Arial")
        ws_stats.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws_stats.cell(row=row, column=2).border = border
        ws_stats.cell(row=row, column=3, value=f"=B{row}/SUM(B6:B{row+len(counts)-1})").font = Font(name="Arial")
        ws_stats.cell(row=row, column=3).number_format = "0.0%"
        ws_stats.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws_stats.cell(row=row, column=3).border = border
        row += 1

    ws_stats.cell(row=row, column=1, value="TOTAL").font = Font(bold=True, name="Arial")
    ws_stats.cell(row=row, column=2, value=f"=SUM(B6:B{row-1})").font = Font(bold=True, name="Arial")
    ws_stats.cell(row=row, column=2).alignment = Alignment(horizontal="center")
    ws_stats.cell(row=row, column=3, value="100%").font = Font(bold=True, name="Arial")
    ws_stats.cell(row=row, column=3).alignment = Alignment(horizontal="center")

    ws_stats.column_dimensions["A"].width = 25
    ws_stats.column_dimensions["B"].width = 12
    ws_stats.column_dimensions["C"].width = 14


# ───────────────────────────────────────────────
# MAIN
# ───────────────────────────────────────────────
def run_scraper(france_travail_id=None, france_travail_secret=None, output_file=OUTPUT_FILE):
    print(f"\n{'='*55}")
    print(f"🚀 Démarrage scraping — {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"{'='*55}\n")

    all_offres = []
    all_offres += scrape_france_travail(france_travail_id, france_travail_secret)
    time.sleep(2)
    all_offres += scrape_wttj()
    time.sleep(2)
    all_offres += scrape_jobteaser()
    time.sleep(2)
    all_offres += scrape_indeed()
    time.sleep(2)
    all_offres += scrape_hellowork()

    all_offres = deduplicate(all_offres)
    all_offres = sort_by_date(all_offres)

    print(f"\n📋 Total après déduplication et tri : {len(all_offres)} offres")
    save_to_excel(all_offres, output_file)
    return len(all_offres)


if __name__ == "__main__":
    FT_CLIENT_ID = os.getenv("FRANCE_TRAVAIL_ID", "")
    FT_CLIENT_SECRET = os.getenv("FRANCE_TRAVAIL_SECRET", "")
    run_scraper(france_travail_id=FT_CLIENT_ID, france_travail_secret=FT_CLIENT_SECRET)
